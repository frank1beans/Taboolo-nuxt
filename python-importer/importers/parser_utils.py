from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from importers.common import (
    _calculate_line_amount,
    _ceil_amount,
    _normalize_wbs7_code,
    _looks_like_wbs7_code,
)
from importers.excel.types import ParsedEstimate, ParsedItem, ParsedWbsLevel


def parse_custom_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
    *,
    combine_totals: bool = True,
) -> ParsedEstimate:
    """Public API: wrapper around `_parse_custom_return_excel`."""
    return _parse_custom_return_excel(
        file_path,
        sheet_name,
        code_columns,
        description_columns,
        price_column,
        quantity_column,
        progressive_column,
        combine_totals=combine_totals,
    )


def _parse_custom_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
    *,
    combine_totals: bool = True,
) -> ParsedEstimate:
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        sheet = _select_sheet(workbook, sheet_name)
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    # Pre-normalize table: drop fully empty columns and fill missing values where needed
    rows, dropped_columns, kept_column_indexes = _drop_empty_columns(raw_rows)
    header_idx = _locate_header_row(rows)
    if header_idx is None:
        raise ValueError("The selected Excel sheet does not contain importable rows.")
    data_rows = rows[header_idx + 1 :]
    if not data_rows:
        raise ValueError("The selected sheet does not contain data after the header row.")

    header_row = rows[header_idx]
    detection = _detect_column_suggestions(rows, header_idx)
    suggestions: dict[str, _ColumnSuggestion] = detection.get("suggestions") if detection else {}
    profiles: list[_ColumnProfile] = detection.get("profiles") if detection else []

    def _format_profiles() -> str:
        entries: list[str] = []
        for profile in profiles[:10]:
            label = profile.header_label or "?"
            entries.append(f"{profile.letter}: {label}")
        return ", ".join(entries)

    def _warn_and_use(target: str, suggestion: _ColumnSuggestion | None, warnings: list[str]) -> int | None:
        if suggestion is None:
            return None
        label = suggestion.header_label or suggestion.column_letter
        warnings.append(
            f"Column {target} not found in saved configuration. "
            f"Automatically using '{label}' ({suggestion.column_letter})."
        )
        return suggestion.column_index

    def _ensure_indexes(name: str, indexes: list[int], warnings: list[str]) -> list[int]:
        if indexes and _column_has_values(data_rows, indexes):
            return indexes
        suggestion = suggestions.get(name)
        fallback = _warn_and_use(name, suggestion, warnings)
        if fallback is not None:
            return [fallback]
        available = _format_profiles()
        raise ValueError(
            f"Cannot find column {name}. "
            f"Detected headers: {available or 'no valid headers.'}"
        )

    column_warnings: list[str] = []
    if dropped_columns:
        column_warnings.append(
            f"Automatically ignored {dropped_columns} completely empty columns."
        )
    try:
        code_indexes = _columns_to_indexes(code_columns, "codice", header_row=header_row, required=False)
    except ValueError:
        code_indexes = []
    try:
        description_indexes = _columns_to_indexes(description_columns, "descrizione", header_row=header_row, required=False)
    except ValueError:
        description_indexes = []
    try:
        price_index = _single_column_index(price_column, "prezzo unitario", header_row=header_row)
    except ValueError:
        price_index = None
    try:
        quantity_index = _single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    except ValueError:
        quantity_index = None
    try:
        progressive_index = _single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)
    except ValueError:
        progressive_index = None

    code_indexes = _ensure_indexes("codice", code_indexes, column_warnings)
    description_indexes = _ensure_indexes("descrizione", description_indexes, column_warnings)
    if price_index is None:
        suggestion = suggestions.get("prezzo")
        price_index = _warn_and_use("prezzo", suggestion, column_warnings)
    if price_index is None:
        available = _format_profiles()
        raise ValueError(
            "Cannot find the unit price column. "
            f"Detected headers: {available or 'no valid headers.'}"
        )
    if not _column_has_values(data_rows, [price_index]):
        raise ValueError(
            f"The selected price column ({get_column_letter(price_index + 1)}) is empty. "
            "Check the chosen column in the sheet."
        )
    if quantity_index is None:
        quantity_index = _warn_and_use("quantity", suggestions.get("quantita"), column_warnings)
    elif quantity_index is not None and not _column_has_values(data_rows, [quantity_index]):
        raise ValueError(
            f"The selected quantity column ({get_column_letter(quantity_index + 1)}) is empty. "
            "Check the chosen column in the sheet."
        )
    if progressive_index is None:
        progressive_index = _warn_and_use("progressivo", suggestions.get("progressivo"), column_warnings)
    elif progressive_index is not None and not _column_has_values(data_rows, [progressive_index]):
        raise ValueError(
            f"The selected progressive column ({get_column_letter(progressive_index + 1)}) is empty. "
            "Check the chosen column in the sheet."
        )
    if not code_indexes and not description_indexes and progressive_index is None:
        raise ValueError(
            "Select at least one column for code, description, or progressive."
        )
    workbook_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    try:
        formula_sheet = _select_sheet(workbook_formulas, sheet_name)
        raw_formula_rows = list(formula_sheet.iter_rows(min_row=header_idx + 2, values_only=False))
        formula_rows = _apply_column_filter(raw_formula_rows, kept_column_indexes)
    finally:
        workbook_formulas.close()

    formula_rows_iter = iter(formula_rows)

    items: list[ParsedItem] = []
    order = 0

    # Head-to-tail logic: 1 progressive = 1 item
    cleaned_rows = []
    for row in data_rows:
        if not _row_has_values(row):
            continue

        code_value = _combine_code(row, code_indexes)
        description_value = _combine_text(row, description_indexes)
        raw_price = _cell_to_float(row, price_index)
        quantity_value_head = _cell_to_float(row, quantity_index) if quantity_index is not None else None
        progressive_value = _cell_to_progressive(row, progressive_index)

        # Skip summary rows only if empty
        if (
            description_value
            and description_value.lower().startswith("totale ")
            and not progressive_value
            and raw_price is None
            and quantity_value_head is None
        ):
            continue

        # Skip chapter/title rows
        if (code_value or description_value) and progressive_value is None and quantity_value_head is None and raw_price is None:
            continue

        cleaned_rows.append(row)

    progressive_indexes = []
    for idx, row in enumerate(cleaned_rows):
        progressive_value = _cell_to_progressive(row, progressive_index)
        if progressive_value is not None:
            progressive_indexes.append((idx, progressive_value, row))

    if not progressive_indexes:
        raise ValueError("No progressive value found: select the correct column.")

    for prog_idx, prog_value, row in progressive_indexes:
        description_value = _combine_text(row, description_indexes)
        unit_price = _sanitize_price_candidate(_cell_to_float(row, price_index))

        # Quantity: either from the same row or from following measure rows
        quantity_value = _cell_to_float(row, quantity_index) if quantity_index is not None else None
        if quantity_value is None:
            quantity_value = _collect_measures(cleaned_rows, prog_idx)

        if quantity_value is None:
            logger_warning = (
                f"Progressive {prog_value}: missing quantity, defaulting to 0."
            )
            print(logger_warning)
            quantity_value = 0.0

        quantity_value, amount = _calculate_line_amount(quantity_value, unit_price)

        item = ParsedItem(
            order=order,
            progressive=prog_value,
            code=_combine_code(row, code_indexes),
            description=description_value,
            wbs_levels=_extract_wbs_levels(row, description_indexes),
            unit=None,
            quantity=quantity_value,
            unit_price=unit_price,
            amount=amount,
            notes=None,
            metadata={
                "sheet_name": sheet.title if sheet else None,
                "column_warnings": column_warnings or None,
            },
        )
        items.append(item)
        order += 1

    # If combine_totals=False, skip sum of total rows
    total_amount = sum([item.amount or 0 for item in items]) if combine_totals else None
    return ParsedEstimate(
        title=sheet.title if sheet else None,
        total_amount=total_amount,
        total_quantity=None,
        items=items,
        stats=None,
    )

# ---- Helper e classi interne (copiati/adattati) ----

def _select_sheet(workbook, sheet_name: str | None):
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.active


def _drop_empty_columns(rows: list[tuple[Any, ...]]):
    if not rows:
        return [], 0, []
    df = pd.DataFrame(rows)
    keep_mask = ~(df.isna().all(axis=0))
    kept_indexes = [idx for idx, keep in enumerate(keep_mask) if keep]
    dropped = len(rows[0]) - len(kept_indexes)
    cleaned = [tuple(val for idx, val in enumerate(row) if keep_mask.iloc[idx]) for row in rows]
    return cleaned, dropped, kept_indexes


def _locate_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for idx, row in enumerate(rows[:10]):
        values = [cell for cell in row if cell is not None]
        if len(values) >= 2:
            return idx
    return None


def _detect_column_suggestions(rows: list[tuple[Any, ...]], header_idx: int):
    if header_idx < 0 or header_idx >= len(rows):
        return {"suggestions": {}, "profiles": []}
    header = rows[header_idx]
    profiles: list[_ColumnProfile] = []
    suggestions: dict[str, _ColumnSuggestion] = {}
    for idx, value in enumerate(header):
        if value is None:
            continue
        label = str(value).strip()
        letter = get_column_letter(idx + 1)
        profile = _ColumnProfile(column_index=idx, column_letter=letter, header_label=label)
        profiles.append(profile)
        normalized = label.lower()
        if "cod" in normalized:
            suggestions.setdefault("codice", _ColumnSuggestion(idx, letter, label))
        if "desc" in normalized or "indicazione" in normalized:
            suggestions.setdefault("descrizione", _ColumnSuggestion(idx, letter, label))
        if "prezzo" in normalized:
            suggestions.setdefault("prezzo", _ColumnSuggestion(idx, letter, label))
        if "quant" in normalized:
            suggestions.setdefault("quantita", _ColumnSuggestion(idx, letter, label))
        if "prog" in normalized:
            suggestions.setdefault("progressivo", _ColumnSuggestion(idx, letter, label))
    return {"suggestions": suggestions, "profiles": profiles}


def _columns_to_indexes(
    columns: Sequence[str],
    name: str,
    *,
    header_row: Sequence[Any] | None,
    required: bool = True,
) -> list[int]:
    if not columns:
        if required:
            raise ValueError(f"Columns {name} not provided")
        return []
    indexes: list[int] = []
    header_lookup = {str(val).strip().lower(): idx for idx, val in enumerate(header_row or []) if val}
    for col in columns:
        col_clean = col.strip()
        try:
            # If it is a letter (A,B,AA) convert to 0-based index
            index = column_index_from_string(col_clean) - 1
        except Exception:
            # Try matching the header title
            index = header_lookup.get(col_clean.lower())
        if index is None:
            if required:
                raise ValueError(f"Column {name} '{col}' not found")
            continue
        indexes.append(index)
    if not indexes and required:
        raise ValueError(f"Column {name} not found")
    return indexes


def _single_column_index(
    column: str | None,
    name: str,
    *,
    header_row: Sequence[Any] | None,
    required: bool = True,
) -> int | None:
    if column is None and not required:
        return None
    if column is None and required:
        raise ValueError(f"Column {name} not provided")
    indexes = _columns_to_indexes([column or ""], name, header_row=header_row, required=required)
    return indexes[0] if indexes else None


def _row_has_values(row: tuple[Any, ...]) -> bool:
    return any(cell not in (None, "", " ") for cell in row)


def _column_has_values(rows: list[tuple[Any, ...]], indexes: list[int]) -> bool:
    for row in rows:
        for idx in indexes:
            if idx < len(row) and row[idx] not in (None, "", " "):
                return True
    return False


def _combine_code(row: tuple[Any, ...], indexes: list[int]) -> str | None:
    parts = []
    for idx in indexes:
        if idx < len(row):
            value = row[idx]
            if value:
                parts.append(str(value).strip())
    return " ".join(parts) if parts else None


def _combine_text(row: tuple[Any, ...], indexes: list[int]) -> str | None:
    parts = []
    for idx in indexes:
        if idx < len(row):
            value = row[idx]
            if value:
                parts.append(str(value).strip())
    return " ".join(parts) if parts else None


def _cell_to_float(row: tuple[Any, ...], index: int | None) -> float | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _cell_to_progressive(row: tuple[Any, ...], index: int | None) -> int | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _apply_column_filter(rows: list[tuple[Any, ...]], indexes: list[int]):
    filtered: list[tuple[Any, ...]] = []
    for row in rows:
        filtered.append(tuple(val for idx, val in enumerate(row) if idx in indexes))
    return filtered


def _extract_wbs_levels(row: tuple[Any, ...], description_indexes: list[int]) -> list[ParsedWbsLevel]:
    levels: list[ParsedWbsLevel] = []
    # Try to infer WBS7 from description/code
    for idx in description_indexes:
        if idx < len(row):
            value = row[idx]
            if value and isinstance(value, str) and _looks_like_wbs7_code(value):
                normalized = _normalize_wbs7_code(value)
                if normalized:
                    levels.append(ParsedWbsLevel(level=7, code=normalized, description=None))
                    break
    return levels


def _collect_measures(rows: list[tuple[Any, ...]], start_idx: int) -> float | None:
    total: float | None = None
    for row in rows[start_idx + 1 :]:
        if not _row_has_values(row):
            continue
        maybe_price = _cell_to_float(row, 0)
        maybe_qty = _cell_to_float(row, 1)
        if maybe_qty is not None and maybe_price is None:
            if total is None:
                total = 0.0
            total += maybe_qty
        else:
            break
    return total


def _sanitize_price_candidate(value: float | None) -> float | None:
    if value is None:
        return None
    if value < 0:
        return None
    return _ceil_amount(value)


def _has_external_formula(row: tuple[Any, ...] | None, index: int | None) -> bool:
    """
    Detect if the cell at given index contains an external formula reference.
    External references usually contain '!' or '[' in the formula string.
    """
    if row is None or index is None:
        return False
    if index >= len(row):
        return False
    cell = row[index]
    try:
        value = cell.value  # type: ignore[attr-defined]
    except Exception:
        value = None
    if not isinstance(value, str):
        return False
    if not value.startswith("="):
        return False
    text = value.lower()
    return "!" in text or "[" in text


@dataclass
class _ColumnProfile:
    column_index: int
    column_letter: str
    header_label: str | None = None


@dataclass
class _ColumnSuggestion:
    column_index: int
    column_letter: str
    header_label: str | None = None


# Narrow export kept for backward compatibility with old importers
_CustomReturnParseResult = ParsedEstimate
