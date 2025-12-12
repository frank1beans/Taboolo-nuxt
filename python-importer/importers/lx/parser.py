from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from importers.common import _calculate_line_amount, _ceil_amount
from importers.excel.types import ParsedEstimate, ParsedItem, ParsedWbsLevel
from importers.helpers.text_and_measure import head_to_tail_quantity, tokenize_description
from importers.parser_utils import (
    _CustomReturnParseResult,
    _apply_column_filter,
    _cell_to_float,
    _columns_to_indexes,
    _combine_code,
    _combine_text,
    _detect_column_suggestions,
    _drop_empty_columns,
    _has_external_formula,
    _locate_header_row,
    _looks_like_wbs7_code,
    _normalize_wbs7_code,
    _sanitize_price_candidate,
    _select_sheet,
    _single_column_index,
)


def parse_lx_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
) -> _CustomReturnParseResult:
    """
    LX linear parser: one row = one item, no header/total combination.
    """
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        sheet = _select_sheet(workbook, sheet_name)
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    rows, dropped_columns, kept_column_indexes = _drop_empty_columns(raw_rows)
    header_idx = _locate_header_row(rows)
    if header_idx is None:
        raise ValueError("Il foglio Excel selezionato non contiene righe valide da importare")
    data_rows = rows[header_idx + 1 :]
    if not data_rows:
        raise ValueError("Il foglio Excel selezionato non contiene dati dopo l'intestazione")

    header_row = rows[header_idx]
    detection = _detect_column_suggestions(rows, header_idx)
    suggestions = detection.get("suggestions") if detection else {}
    column_warnings: list[str] = []
    if dropped_columns:
        column_warnings.append(f"Ignorate automaticamente {dropped_columns} colonne completamente vuote.")

    code_indexes = _ensure_indexes_lc("codice", code_columns, data_rows, header_row, column_warnings, suggestions)
    description_indexes = _ensure_indexes_lc("descrizione", description_columns, data_rows, header_row, column_warnings, suggestions)

    price_index = _single_column_index(price_column, "prezzo unitario", header_row=header_row)
    if not _has_values(data_rows, [price_index]):
        raise ValueError(
            f"La colonna prezzo selezionata ({get_column_letter(price_index + 1)}) non contiene valori. "
            "Verifica la colonna scelta nel foglio."
        )

    quantity_index = _single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    progressive_index = _single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)

    workbook_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    try:
        formula_sheet = _select_sheet(workbook_formulas, sheet_name)
        raw_formula_rows = list(formula_sheet.iter_rows(min_row=header_idx + 2, values_only=False))
        formula_rows = _apply_column_filter(raw_formula_rows, kept_column_indexes)
    finally:
        workbook_formulas.close()

    voci: list[ParsedItem] = []
    ordine = 0
    for data_row, formula_row in zip(data_rows, formula_rows):
        codice = _combine_code(data_row, code_indexes)
        descrizione = _combine_text(data_row, description_indexes)
        progressivo = _cell_to_progressive(data_row, progressive_index)
        quantita = _cell_to_float(data_row, quantity_index) if quantity_index is not None else None
        prezzo_unitario = _sanitize_price_candidate(_cell_to_float(data_row, price_index))
        has_formula = _has_external_formula(formula_row, quantity_index)
        tokens = tokenize_description(descrizione or "")

        # Skip item if quantity is an external formula or missing
        if has_formula or quantita is None:
            continue

        quantita, importo = _calculate_line_amount(quantita, prezzo_unitario)

        voce = ParsedItem(
            order=ordine,
            progressive=progressivo,
            code=codice,
            description=descrizione,
            wbs_levels=_extract_wbs_levels(data_row, description_indexes),
            unit=None,
            quantity=quantita,
            unit_price=prezzo_unitario,
            amount=importo,
            notes=None,
            metadata={
                "column_warnings": column_warnings or None,
                "tokens": tokens,
            },
        )
        voci.append(voce)
        ordine += 1

    total_amount = _ceil_amount(sum([voce.amount or 0 for voce in voci])) if voci else None
    return ParsedEstimate(
        title=sheet.title if sheet else None,
        total_amount=total_amount,
        total_quantity=None,
        items=voci,
        stats=None,
    )


def _ensure_indexes_lc(
    name: str,
    columns: Sequence[str],
    data_rows,
    header_row,
    warnings: list[str],
    suggestions: dict,
) -> list[int]:
    try:
        indexes = _columns_to_indexes(columns, name, header_row=header_row, required=False)
    except ValueError:
        indexes = []
    if indexes and _has_values(data_rows, indexes):
        return indexes
    suggestion = suggestions.get(name)
    if suggestion:
        warnings.append(
            f"Colonna {name} non trovata, uso automaticamente '{suggestion.header_label or suggestion.column_letter}'."
        )
        return [suggestion.column_index]
    available = ", ".join([cell for cell in header_row if cell]) if header_row else ""
    raise ValueError(f"Impossibile individuare la colonna {name}. Intestazioni rilevate: {available or 'nessuna'}")


def _has_values(rows, indexes: list[int]) -> bool:
    for row in rows:
        for idx in indexes:
            if idx < len(row) and row[idx] not in (None, "", " "):
                return True
    return False


def _cell_to_progressive(row, index: int | None) -> int | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _extract_wbs_levels(row, description_indexes: list[int]) -> list[ParsedWbsLevel]:
    levels: list[ParsedWbsLevel] = []
    for idx in description_indexes:
        if idx < len(row):
            value = row[idx]
            if value and isinstance(value, str) and _looks_like_wbs7_code(value):
                normalized = _normalize_wbs7_code(value)
                if normalized:
                    levels.append(ParsedWbsLevel(level=7, code=normalized, description=None))
                    break
    return levels
