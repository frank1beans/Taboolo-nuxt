from __future__ import annotations

import re
import unicodedata
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook

from importers.excel.types import ParsedEstimate, ParsedItem, ParsedWbsLevel


MAX_WBS_LEVELS = 7

_WBS6_PATTERN = re.compile(r"^[A-Za-z]\d{3}$")
_WBS7_PATTERN = re.compile(r"^[A-Za-z]\d{3}[.\s_-]?\d{3}$")


def parse_estimate_excel(
    path: Path,
    sheet_name: str | None = None,
    *,
    price_column: str | None = None,
    quantity_column: str | None = None,
) -> ParsedEstimate:
    workbook = load_workbook(path, data_only=True)
    sheet = _pick_sheet(workbook, sheet_name or "")
    if sheet is None:
        raise ValueError("Cannot find the bill-of-quantities sheet")

    ws = workbook[sheet]
    rows = _iter_rows(ws)

    if _is_bill_of_quantities(rows):
        return _parse_bill_of_quantities(ws.title, rows)
    return _parse_estimate(ws.title, rows, price_column=price_column, quantity_column=quantity_column)


# ---------------------------------------------------------------------------
# Bill of quantities / estimate parser
# ---------------------------------------------------------------------------


def _parse_estimate(
    title: str | None,
    rows: list[list],
    *,
    price_column: str | None = None,
    quantity_column: str | None = None,
) -> ParsedEstimate:
    header_row = _find_header_row(rows)
    if header_row is None:
        raise ValueError("Header not recognized: unable to locate main columns")

    # Dynamic column mapping (supports export with progressive in column 0)
    normalized_header = [_normalize_header(cell) for cell in rows[header_row]]

    def _find_col(candidates: set[str]) -> int | None:
        for idx, value in enumerate(normalized_header):
            if not value:
                continue
            tokens = set(value.split())
            if candidates & tokens:
                return idx
        return None

    progressivo_idx = _find_col({"progressivo", "n", "n.", "num"})
    if progressivo_idx is None:
        progressivo_idx = 1
    codice_idx = _find_col({"codice", "cod"})
    if codice_idx is None:
        codice_idx = 2
    descr_idx = _find_col({"descrizione", "indicazione"})
    if descr_idx is None:
        descr_idx = 3
    unita_idx = _find_col({"u.m.", "um", "unita", "unita", "unita"})  # include common unit synonyms
    if unita_idx is None:
        unita_idx = 4
    quantita_idx = None
    normalized_quantity = _normalize_header(quantity_column) if quantity_column else None
    for idx, value in enumerate(normalized_header):
        if normalized_quantity and value == normalized_quantity:
            quantita_idx = idx
            break
        if value and (value.replace(" ", "").startswith("quant") or value in {"qta", "qt", "q"}):
            quantita_idx = idx
            if not normalized_quantity:
                break
    if quantita_idx is None:
        quantita_idx = 9

    def _find_prezzocol(predicate) -> int | None:
        for idx, value in enumerate(normalized_header):
            if value and predicate(value):
                return idx
        return None

    normalized_price = _normalize_header(price_column) if price_column else None
    prezzo_idx = None
    if normalized_price:
        prezzo_idx = _find_prezzocol(lambda v: v == normalized_price)
    if prezzo_idx is None:
        prezzo_idx = _find_prezzocol(lambda v: "prezzo" in v and "importo" not in v) or 10
    importo_idx = _find_prezzocol(lambda v: "importo" in v or v.endswith("totale")) or 11

    current_wbs: list[ParsedWbsLevel | None] = [None] * MAX_WBS_LEVELS
    items: list[ParsedItem] = []
    ordine = 0

    def _row_is_item(row) -> bool:
        prog = row[progressivo_idx] if len(row) > progressivo_idx else None
        code = row[codice_idx] if len(row) > codice_idx else None
        descr = row[descr_idx] if len(row) > descr_idx else None
        if prog in (None, "", " "):
            return False
        return bool(code or descr)

    def _row_is_section(row) -> bool:
        prog = row[progressivo_idx] if len(row) > progressivo_idx else None
        code = row[codice_idx] if len(row) > codice_idx else None
        descr = row[descr_idx] if len(row) > descr_idx else None
        if prog not in (None, "", " "):
            return False
        if not code or not descr:
            return False
        value = str(descr).strip().lower()
        return value != "totale"

    def _row_is_total(row) -> bool:
        target_descr = row[descr_idx] if len(row) > descr_idx else (row[codice_idx] if len(row) > codice_idx else None)
        value = str(target_descr).strip().lower() if target_descr else ""
        return value == "totale"

    i = header_row + 1
    while i < len(rows):
        row = rows[i]
        if _row_is_empty(row):
            i += 1
            continue

        if _row_is_total(row):
            i += 1
            continue

        if _row_is_section(row):
            code = _sanitize_code(row[2])
            description = _extract_description(row[3], code)
            level = _guess_wbs_level(code, current_wbs)
            level_idx = max(0, min(level, MAX_WBS_LEVELS) - 1)
            current_wbs[level_idx] = ParsedWbsLevel(
                level=level, code=code, description=description
            )
            for idx in range(level_idx + 1, MAX_WBS_LEVELS):
                current_wbs[idx] = None
            i += 1
            continue

        if not _row_is_item(row):
            i += 1
            continue

    progressivo = _to_int(row[progressivo_idx] if len(row) > progressivo_idx else None)
    codice = _sanitize_code(row[codice_idx] if len(row) > codice_idx else None)
    descrizione = _sanitize_text(row[descr_idx] if len(row) > descr_idx else None)
    if not codice:
        codice = _generate_fallback_code(progressivo, descrizione, ordine)

        # Debug logging for problematic progressives
        import logging
        logger = logging.getLogger(__name__)
        if progressivo in (10, 20, 180, 220):
            logger.info(f"[PARSER DEBUG] Progressivo {progressivo}: codice={codice}, riga={i}")

        (
            unita_misura,
            quantita,
            prezzo_unitario,
            importo,
            note,
            rows_consumed,
        ) = _collect_measure_rows(
            rows,
            i + 1,
            quantita_idx=quantita_idx,
            prezzo_idx=prezzo_idx,
            importo_idx=importo_idx,
            unita_idx=unita_idx,
        )

        if progressivo in (10, 20, 180, 220):
            logger.info(f"[PARSER DEBUG] Dopo _collect_measure_rows: prog={progressivo}, qta={quantita}, prezzo={prezzo_unitario}, importo={importo}, rows_consumed={rows_consumed}")

        if prezzo_unitario is None:
            # Fallback: cerca prezzo nelle righe successive (es. colonne PU specifiche)
            for look_ahead in range(1, 4):
                if i + look_ahead >= len(rows):
                    break
                candidate_row = rows[i + look_ahead]
                prezzo_unitario = _to_float(
                    candidate_row[prezzo_idx] if len(candidate_row) > prezzo_idx else None,
                    decimals=4,
                )
                if prezzo_unitario is not None:
                    break
        if importo is None:
            for look_ahead in range(1, 4):
                if i + look_ahead >= len(rows):
                    break
                candidate_row = rows[i + look_ahead]
                importo = _to_float(
                    candidate_row[importo_idx] if len(candidate_row) > importo_idx else None,
                    decimals=2,
                )
                if importo is not None:
                    break

        # If the main row already has quantity/price/amount, use them (flattened format)
        if quantita in (None,) and len(row) > quantita_idx:
            quantita = _to_float(row[quantita_idx], decimals=2)
        if prezzo_unitario is None and len(row) > prezzo_idx:
            prezzo_unitario = _to_float(row[prezzo_idx], decimals=4)
        if importo is None and len(row) > importo_idx:
            importo = _to_float(row[importo_idx], decimals=2)
        if importo is None and prezzo_unitario is not None and quantita not in (None, 0):
            importo = round(prezzo_unitario * quantita, 2)
        if prezzo_unitario is None and importo is not None and quantita not in (None, 0):
            prezzo_unitario = round(importo / quantita, 4)

        # Check if this is a "Total" row (price value is the total amount of the group)
        is_totale_row = descrizione and "totale" in descrizione.lower()

        if progressivo in (10, 20, 180, 220):
            logger.info(f"[PARSER DEBUG] Prima logica totale: prog={progressivo}, is_totale_row={is_totale_row}, descrizione='{descrizione}', prezzo={prezzo_unitario}, importo={importo}")

        # FIX: more robust calculation to avoid import/price swap confusion
        if is_totale_row and prezzo_unitario is not None:
            # Per righe "Totale": usa il prezzo come importo totale
            logger.warning(f"[PARSER] Voce prog={progressivo} con descrizione '{descrizione}' contiene 'totale' - interpreto prezzo come importo!")
            importo = round(prezzo_unitario, 2)
            prezzo_unitario = None

        if progressivo in (10, 20, 180, 220):
            logger.info(f"[PARSER DEBUG] Dopo logica totale: prog={progressivo}, prezzo={prezzo_unitario}, importo={importo}")
        elif (
            importo is None
            and quantita not in (None, 0)
            and prezzo_unitario is not None
        ):
            importo = round(prezzo_unitario * quantita, 2)
        elif (
            prezzo_unitario is None
            and quantita not in (None, 0)
            and importo is not None
        ):
            # FIX: Sanity check prima di calcolare prezzo da importo
            # If amount/quantity yields a suspicious price (>100000), fields might be swapped
            calculated_price = importo / quantita
            if calculated_price > 100000:
                # Probabile errore: importo totale al posto del prezzo
                # Proviamo a invertire: usa importo come prezzo e ricalcola importo
                prezzo_unitario = round(importo, 4)
                importo = round(prezzo_unitario * quantita, 2)
            else:
                prezzo_unitario = round(calculated_price, 4)

        livelli_voce = [lvl for lvl in current_wbs if lvl is not None]
        livelli_voce = _ensure_wbs_hierarchy(
            livelli_voce,
            codice,
            descrizione,
        )

        items.append(
            ParsedItem(
                order=ordine,
                progressive=progressivo,
                code=codice,
                description=descrizione,
                wbs_levels=livelli_voce,
                unit=unita_misura,
                quantity=quantita,
                unit_price=prezzo_unitario,
                amount=importo,
                notes=note,
                metadata={
                    "source": "excel",
                    "parser": "estimate",
                    "sheet_title": titolo,
                },
            )
        )

        ordine += 1
        i += rows_consumed + 1

    total_amount = round(sum(item.amount or 0 for item in items), 2) if items else None
    quantity_values = [item.quantity for item in items if item.quantity is not None]
    total_quantity = round(sum(quantity_values), 2) if quantity_values else None

    return ParsedEstimate(
        title=titolo,
        total_amount=total_amount,
        total_quantity=total_quantity,
        items=items,
    )


# ---------------------------------------------------------------------------
# Flattened bill-of-quantities parser
# ---------------------------------------------------------------------------


def _parse_bill_of_quantities(titolo: str | None, rows: list[list]) -> ParsedEstimate:
    header_index = _find_header_row_lista(rows)
    if header_index is None:
        raise ValueError("Bill of quantities header not recognized")

    header_row = rows[header_index]
    header_map = {
        _normalize_header(cell): idx
        for idx, cell in enumerate(header_row)
        if _normalize_header(cell)
    }

    def pick(row: Sequence, *keys: str) -> object | None:
        for key in keys:
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    items: list[ParsedItem] = []
    ordine = 0

    for r in range(header_index + 1, len(rows)):
        row = rows[r]
        if _row_is_empty(row):
            continue

        categoria_code = _sanitize_code(
            pick(row, "cod categorie", "cod categoria", "categoria")
        )
        categoria_desc = _sanitize_text(
            pick(row, "des categorie", "desc categorie", "descrizione categoria")
        )
        sub_code = _sanitize_code(
            pick(row, "cod subcategorie", "cod subcategoria")
        )
        sub_desc = _sanitize_text(
            pick(row, "des subcategorie", "desc subcategorie", "descrizione subcategorie")
        )
        descrizione_estesa = _sanitize_text(
            pick(
                row,
                "estesa",
                "descrizione estesa",
                "descrizione",
                "estesa lavorazione",
            )
        )
        unita = _sanitize_text(
            pick(row, "u m", "um", "unita di misura")
        )
        quantita = _to_float(pick(row, "quantita", "qta"), decimals=2)
        prezzo = _to_float(pick(row, "prezzo", "prezzo unitario"), decimals=4)
        importo = _to_float(pick(row, "importo", "totale", "totale importo"), decimals=2)
        note = _sanitize_text(pick(row, "note", "osservazioni"))

        description = descrizione_estesa or sub_desc
        if not description:
            continue

        livelli: list[ParsedWbsLevel] = []
        if categoria_code or categoria_desc:
            livelli.append(
                ParsedWbsLevel(level=5, code=categoria_code, description=categoria_desc)
            )

        livelli = _ensure_wbs_hierarchy(
            livelli,
            sub_code or categoria_code,
            description,
            level6_description=categoria_desc,
        )

        item_code = sub_code or categoria_code

        # Check if this is a "Total" row (price holds the total amount)
        is_total_row = description and "totale" in description.lower()

        # FIX: Stesso sanity check per lista lavorazioni
        if is_total_row and prezzo is not None:
            # For "Total" rows: use price as total amount
            importo = round(prezzo, 2)
            prezzo = None
        elif importo is None and quantita not in (None, 0) and prezzo is not None:
            importo = round(prezzo * quantita, 2)
        elif prezzo is None and quantita not in (None, 0) and importo is not None:
            calculated_price = importo / quantita
            if calculated_price > 100000:
                # Probabile confusione: importo al posto di prezzo
                prezzo = round(importo, 4)
                importo = round(prezzo * quantita, 2)
            else:
                prezzo = round(calculated_price, 4)

        items.append(
            ParsedItem(
                order=ordine,
                progressive=None,
                code=item_code,
                description=description,
                wbs_levels=livelli,
                unit=unita,
                quantity=quantita,
                unit_price=prezzo,
                amount=importo,
                notes=note,
                metadata={
                    "source": "excel",
                    "parser": "bill_of_quantities",
                    "sheet_title": titolo,
                },
            )
        )
        ordine += 1

    total_amount = round(sum(item.amount or 0 for item in items), 2) if items else None
    quantity_values = [item.quantity for item in items if item.quantity is not None]
    total_quantity = (
        round(sum(quantity_values), 2) if quantity_values else None
    )

    return ParsedEstimate(
        title=titolo,
        total_amount=total_amount,
        total_quantity=total_quantity,
        items=items,
    )


# ---------------------------------------------------------------------------
# WBS helpers
# ---------------------------------------------------------------------------


def _ensure_wbs_hierarchy(
    levels: list[ParsedWbsLevel],
    codice: str | None,
    descrizione: str | None,
    *,
    level6_description: str | None = None,
) -> list[ParsedWbsLevel]:
    """Garantisce che la gerarchia WBS si fermi alla WBS6 (formato A###).
    Tutte le voci con codice tipo A001.010 o A001.010.001 vengono assegnate
    alla WBS6 base (A001).
    """
    levels = list(levels)

    code = _sanitize_code(codice)
    base_code: str | None = None

    if code:
        # If this is an exact WBS6 code (A001)
        if _WBS6_PATTERN.fullmatch(code):
            base_code = code
        else:
            # If this is a sub-level (A001.010, A001.010.001, etc.)
            match = re.match(r"^([A-Za-z]\d{3})[.\s_-]", code)
            if match:
                base_code = match.group(1)
            else:
                # Se non riconosciuto, prova a isolare prefisso plausibile
                head = code.split(".")[0]
                if _WBS6_PATTERN.fullmatch(head):
                    base_code = head

    level6 = next((lvl for lvl in levels if lvl.level == 6), None)

    # Pulizia descrizione
    level6_description = _sanitize_level6_description(
        level6_description or descrizione,
        base_code,
    )

    # Aggiorna o crea livello 6
    if base_code:
        if level6 is None:
            levels.append(
                ParsedWbsLevel(
                    level=6,
                    code=base_code,
                    description=level6_description,
                )
            )
        else:
            if not level6.code:
                level6.code = base_code
            if not level6.description and level6_description:
                level6.description = level6_description
    elif level6_description:
        if level6 is None:
            levels.append(
                ParsedWbsLevel(level=6, code=None, description=level6_description)
            )
        elif not level6.description:
            level6.description = level6_description

    # Ensure we do not exceed level 6
    levels = [lvl for lvl in levels if lvl.level <= 6]
    levels.sort(key=lambda lvl: lvl.level)
    return levels

def _sanitize_level6_description(
    description: str | None,
    base_code: str | None,
) -> str | None:
    if not description:
        return None
    cleaned = description.strip()
    if base_code:
        pattern = re.compile(rf"^\s*{re.escape(base_code)}\s*[-–:]\s*", re.IGNORECASE)
        cleaned = pattern.sub("", cleaned)
    cleaned = cleaned.strip()
    return cleaned or None


# ---------------------------------------------------------------------------
# Helper comuni
# ---------------------------------------------------------------------------


def _pick_sheet(workbook, requested: str) -> str | None:
    sheetnames = workbook.sheetnames
    if requested and requested in sheetnames:
        return requested

    # 1️⃣ Keyword-based detection
    preferred_keywords = ("computo", "ritorno", "offerta", "lista")
    for keyword in preferred_keywords:
        for name in sheetnames:
            if keyword in name.lower():
                return name

    # 2️⃣ Detect recognizable headers
    for name in sheetnames:
        ws = workbook[name]
        sample_rows = _iter_rows(ws, max_rows=40)
        if _find_header_row(sample_rows) is not None or _find_header_row_lista(sample_rows) is not None:
            return name

    # 3️⃣ Fallback: pick the sheet with most non-empty cells
    max_nonempty = 0
    best_sheet = None
    for name in sheetnames:
        ws = workbook[name]
        count = sum(
            1
            for row in ws.iter_rows(values_only=True)
            for cell in row
            if cell not in (None, "", " ")
        )
        if count > max_nonempty:
            max_nonempty = count
            best_sheet = name

    return best_sheet or (sheetnames[0] if sheetnames else None)


def _iter_rows(ws, max_rows: int | None = None) -> list[list]:
    rows: list[list] = []
    for index, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if max_rows is not None and index + 1 >= max_rows:
            break
    return rows


def _is_bill_of_quantities(rows: list[list]) -> bool:
    for row in rows[:10]:
        for cell in row:
            if isinstance(cell, str) and "cod. categorie" in cell.lower():
                return True
    return False


def _find_header_row(rows: list[list]) -> int | None:
    for idx, row in enumerate(rows):
        normalized = [_normalize_header(cell) for cell in row]
        if not any(normalized):
            continue

        tokens_sets = [set(value.split()) for value in normalized if value]
        flattened: set[str] = set().union(*tokens_sets) if tokens_sets else set()

        has_code = any(value in {"codice", "cod"} for value in flattened)
        has_description = "descrizione" in flattened
        has_quantity = any(
            value.replace(" ", "").startswith("quant")
            or value.replace(" ", "") in {"qta", "qt", "q"}
            for value in normalized
            if value
        )
        has_price = any("prezzo" in value for value in normalized if value)
        has_amount = any(
            "importo" in value or value.endswith("totale") for value in normalized if value
        )

        if has_code and (
            has_description
            or (
                (has_quantity and has_price)
                or (has_quantity and has_amount)
                or (has_price and has_amount)
            )
        ):
            return idx
    return None


def _find_header_row_lista(rows: list[list]) -> int | None:
    for idx, row in enumerate(rows):
        if row and any(
            isinstance(cell, str) and cell.strip().lower() == "cod. categorie"
            for cell in row
        ):
            return idx
    return None


def _normalize_header(value) -> str:
    if not isinstance(value, str):
        return ""
    normalized = unicodedata.normalize("NFKD", value)
    cleaned = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in normalized)
    cleaned = " ".join(cleaned.strip().lower().split())
    return cleaned


def _row_is_empty(row: Sequence) -> bool:
    return all(cell in (None, "", " ") for cell in row)


def _is_total_row(row: Sequence) -> bool:
    value = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ""
    return value == "totale"


def _is_section_row(row: Sequence) -> bool:
    codice = row[2]
    descrizione = row[3]
    progressivo = row[1]
    if progressivo not in (None, "", " "):
        return False
    if not codice or not descrizione:
        return False
    value = str(descrizione).strip().lower()
    return value != "totale"


def _is_item_row(row: Sequence) -> bool:
    progressivo = row[1]
    codice = row[2]
    descrizione = row[3]
    if progressivo in (None, "", " "):
        return False
    if not codice and not descrizione:
        return False
    return True


def _sanitize_code(value) -> str | None:
    text = _sanitize_text(value)
    return text.strip() if text else None


def _generate_fallback_code(
    progressivo: int | None, descrizione: str | None, ordine: int
) -> str:
    if progressivo is not None:
        return f"PROG-{int(progressivo):05d}"
    if descrizione:
        slug = _slugify(descrizione)
        if slug:
            return f"DESC-{slug}"
    return f"ORD-{ordine:05d}"


def _extract_description(cell_value, code: str | None) -> str | None:
    text = _sanitize_text(cell_value)
    if not text:
        return None
    if code:
        prefix = f"{code} - "
        if text.startswith(prefix):
            return text[len(prefix) :].strip()
        if text.startswith(code):
            remainder = text[len(code) :].lstrip(" -")
            return remainder or text
    return text


def _guess_wbs_level(code: str | None, current: Sequence[ParsedWbsLevel | None]) -> int:
    if not code:
        for idx, item in enumerate(current):
            if item is None:
                return idx + 1
        return MAX_WBS_LEVELS

    core = code.strip()
    segments = core.split(".")
    if len(segments) >= 2 and _WBS6_PATTERN.match(segments[0]):
        return 7
    if len(segments) >= 3:
        return 7
    if _WBS7_PATTERN.match(core):
        return 7
    if _WBS6_PATTERN.match(core):
        return 6
    if core.isdigit():
        return 1
    if len(core) == 1 and core.isalpha():
        return 2
    if core.isalpha():
        if len(core) <= 3:
            return 3
    if len(core) == 3 and core[0].isalpha() and core[1:].isdigit():
        return 4
    if core[0].isalpha() and core[1:].isdigit():
        return 5

    for idx, item in enumerate(current):
        if item is None:
            return idx + 1
    return MAX_WBS_LEVELS


def _collect_measure_rows(
    rows: list[list],
    start_index: int,
    *,
    quantita_idx: int = 9,
    prezzo_idx: int = 10,
    importo_idx: int = 11,
    unita_idx: int | None = None,
    progressivo_idx: int = 1,
    codice_idx: int = 2,
    descr_idx: int = 3,
) -> tuple[str | None, float | None, float | None, float | None, str | None, int]:
    unita: str | None = None
    quantita_total: float | None = None
    prezzo: float | None = None
    importo_total: float | None = None
    note: str | None = None
    rows_consumed = 0

    for offset in range(start_index, len(rows)):
        current = rows[offset]
        if _row_is_empty(current):
            rows_consumed += 1
            continue

        prog = current[progressivo_idx] if len(current) > progressivo_idx else None
        code = current[codice_idx] if len(current) > codice_idx else None
        descr = current[descr_idx] if len(current) > descr_idx else None
        if (prog not in (None, "", " ") and (code or descr)) or (
            prog in (None, "", " ") and code and descr and str(descr).strip().lower() != "totale"
        ):
            break

        raw_descr = current[descr_idx] if len(current) > descr_idx else None
        normalized_descr = (
            str(raw_descr).strip().lower() if raw_descr is not None else ""
        )
        value_descr = normalized_descr
        current_code = _sanitize_code(current[2] if len(current) > 2 else None)
        progressivo = current[1] if len(current) > 1 else None

        if (
            progressivo in (None, "", " ")
            and not current_code
            and normalized_descr.startswith("totale ")
            and normalized_descr != "totale"
        ):
            rows_consumed += 1
            break

        # Rileva righe di detrazione: se la descrizione contiene "detraz" consideriamo la misura negativa.
        detrazione = _row_has_detrazione(current)
        sign = -1 if detrazione else 1
        is_total_row = value_descr.startswith("totale")

        unita_candidate = _sanitize_text(current[unita_idx]) if unita_idx is not None and len(current) > unita_idx else _sanitize_text(current[4] if len(current) > 4 else None)
        if unita_candidate:
            unita = unita_candidate

        quantita_candidate = _to_float(current[quantita_idx] if len(current) > quantita_idx else None, decimals=2)
        if quantita_candidate is not None:
            # Use the last non-None value instead of summing (avoids double counting when Total row repeats quantity)
            quantita_total = sign * quantita_candidate

        prezzo_candidate = _to_float(current[prezzo_idx] if len(current) > prezzo_idx else None, decimals=4)
        if prezzo_candidate is None and len(current) > prezzo_idx:
            # Cerca un prezzo nelle colonne successive (es. PU per impresa specifica)
            for alt_idx in range(prezzo_idx + 1, min(len(current), importo_idx)):
                prezzo_candidate = _to_float(current[alt_idx], decimals=4)
                if prezzo_candidate is not None:
                    break
        if prezzo_candidate is not None:
            prezzo = prezzo_candidate

        importo_candidate = _to_float(current[importo_idx] if len(current) > importo_idx else None, decimals=2)
        if importo_candidate is not None:
            importo_total = sign * importo_candidate

        if len(current) > 12:
            note_candidate = _sanitize_text(current[12])
            if note_candidate:
                note = note_candidate

        rows_consumed += 1

        if is_total_row:
            break

    quantita_value: float | None = None
    importo_value: float | None = None
    if quantita_total is not None:
        quantita_value = round(quantita_total, 2)
    if importo_total is not None:
        importo_value = round(importo_total, 2)

    # Se manca l'importo ma abbiamo prezzo e quantita, calcola l'importo; viceversa calcola il prezzo.
    if importo_value is None and prezzo is not None and quantita_value not in (None, 0):
        importo_value = round(prezzo * quantita_value, 2)
    elif prezzo is None and quantita_value not in (None, 0) and importo_value is not None:
        prezzo = round(importo_value / quantita_value, 4)

    return unita, quantita_value, prezzo, importo_value, note, rows_consumed


def _sanitize_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("_x000D_", "\n").strip()
    return text or None


def _row_has_detrazione(row: Sequence) -> bool:
    """Riconosce righe di misura marcate come detrazioni."""
    for cell in row:
        if isinstance(cell, str) and "detraz" in cell.lower():
            return True
    return False


def _to_float(value, decimals: int = 2) -> float | None:
    """Converte un valore in float, arrotondando al numero di decimali specificato."""
    if value is None or value == "":
        return None

    quantize_str = "0." + "0" * decimals if decimals > 0 else "1"

    if isinstance(value, (int, float)):
        try:
            result = Decimal(str(value)).quantize(
                Decimal(quantize_str), rounding=ROUND_HALF_UP
            )
            return float(result)
        except Exception:
            return None
    text = str(value).strip().replace("\u202f", "")
    text = text.replace(".", "").replace(",", ".")
    try:
        result = Decimal(text).quantize(
            Decimal(quantize_str), rounding=ROUND_HALF_UP
        )
        return float(result)
    except Exception:
        return None


def _to_int(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    cleaned = "".join(
        ch for ch in normalized.lower() if ch.isalnum()
    )
    return cleaned[:16]
