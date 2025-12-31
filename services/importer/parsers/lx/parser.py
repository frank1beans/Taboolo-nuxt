"""
LX return file parser.
Depends on base_types, math_utils, wbs_utils, and excel utilities.
"""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from parsers.shared.base_types import ParsedEstimate, ParsedItem, ParsedWbsLevel
from parsers.shared.math_utils import calculate_line_amount, ceil_amount
from parsers.shared.wbs_utils import normalize_wbs7_code, looks_like_wbs7_code
from parsers.helpers.text_and_measure import head_to_tail_quantity, tokenize_description
from parsers.excel.detector import (
    detect_column_suggestions,
    columns_to_indexes,
    single_column_index,
)
from parsers.excel.processor import (
    drop_empty_columns,
    locate_header_row,
    column_has_values,
    combine_code,
    combine_text,
    cell_to_float,
    apply_column_filter,
    has_external_formula,
    sanitize_price_candidate,
)
from parsers.excel.reader import select_sheet


def parse_lx_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
    header_row_index: int | None = None,
    long_description_columns: Sequence[str] | None = None,
) -> ParsedEstimate:
    """
    LX linear parser: one row = one item, no header/total combination.
    """
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    try:
        sheet = select_sheet(workbook, sheet_name)
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    rows, dropped_columns, kept_column_indexes = drop_empty_columns(raw_rows)

    if header_row_index is not None and header_row_index >= 0:
        header_idx = header_row_index
    else:
        header_idx = locate_header_row(rows)

    if header_idx is None:
        raise ValueError("Il foglio Excel selezionato non contiene righe valide da importare")
    data_rows = rows[header_idx + 1:]
    if not data_rows:
        raise ValueError("Il foglio Excel selezionato non contiene dati dopo l'intestazione")

    header_row = rows[header_idx]
    detection = detect_column_suggestions(rows, header_idx)
    suggestions = detection.get("suggestions") if detection else {}
    column_warnings: list[str] = []
    if dropped_columns:
        column_warnings.append(f"Ignorate automaticamente {dropped_columns} colonne completamente vuote.")

    # Try to resolve code columns - optional if description is provided
    # If the user did not select a code column, do NOT auto-suggest one: ignore codes entirely.
    try:
        if code_columns and any(col for col in code_columns):
            code_indexes = _ensure_indexes_lc("codice", code_columns, data_rows, header_row, column_warnings, suggestions)
        else:
            code_indexes = []
    except ValueError:
        code_indexes = []  # Code is optional if description is present
    
    # Try to resolve description columns - optional if code is provided
    try:
        description_indexes = _ensure_indexes_lc("descrizione", description_columns, data_rows, header_row, column_warnings, suggestions)
    except ValueError:
        description_indexes = []  # Description is optional if code is present

    # Try to resolve long description columns - optional
    try:
        if long_description_columns and any(col for col in long_description_columns):
            long_description_indexes = _ensure_indexes_lc("descrizione_estesa", long_description_columns, data_rows, header_row, column_warnings, suggestions)
        else:
            long_description_indexes = []
    except ValueError:
        long_description_indexes = []
    
    # At least one of code or description or long description must be present
    if not code_indexes and not description_indexes and not long_description_indexes:
        available = ", ".join([str(cell) for cell in header_row if cell]) if header_row else ""
        raise ValueError(
            f"È necessario selezionare almeno una colonna codice, descrizione o descrizione estesa. "
            f"Intestazioni rilevate: {available or 'nessuna'}"
        )

    price_index = single_column_index(price_column, "prezzo unitario", header_row=header_row)
    if not _has_values(data_rows, [price_index]):
        raise ValueError(
            f"La colonna prezzo selezionata ({get_column_letter(price_index + 1)}) non contiene valori. "
            "Verifica la colonna scelta nel foglio."
        )

    quantity_index = single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    progressive_index = single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)

    workbook_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    try:
        formula_sheet = select_sheet(workbook_formulas, sheet_name)
        raw_formula_rows = list(formula_sheet.iter_rows(min_row=header_idx + 2, values_only=False))
        formula_rows = apply_column_filter(raw_formula_rows, kept_column_indexes)
    finally:
        workbook_formulas.close()

    # Debug logging
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"[LX Parser] Header row index: {header_idx}")
    logger.info(f"[LX Parser] Header content: {header_row[:10] if header_row else 'None'}...")
    logger.info(f"[LX Parser] Code indexes: {code_indexes}, Description indexes: {description_indexes}")
    logger.info(f"[LX Parser] Long Description indexes: {long_description_indexes}")
    logger.info(f"[LX Parser] Price index: {price_index}, Quantity index: {quantity_index}")
    logger.info(f"[LX Parser] Data rows count: {len(data_rows)}")
    if data_rows:
        logger.info(f"[LX Parser] First data row: {data_rows[0][:10]}...")

    voci: list[ParsedItem] = []
    ordine = 0
    skipped_reasons = {"no_quantity": 0, "has_formula": 0}
    for data_row, formula_row in zip(data_rows, formula_rows):
        codice = combine_code(data_row, code_indexes)
        descrizione = combine_text(data_row, description_indexes)
        descrizione_estesa = combine_text(data_row, long_description_indexes)
        progressivo = _cell_to_progressive(data_row, progressive_index)
        quantita = cell_to_float(data_row, quantity_index) if quantity_index is not None else None
        prezzo_unitario = sanitize_price_candidate(cell_to_float(data_row, price_index))
        has_formula = has_external_formula(formula_row, quantity_index)
        tokens = tokenize_description(descrizione or "")

        # Skip item if quantity is an external formula or missing
        if has_formula:
            skipped_reasons["has_formula"] += 1
            continue
        if quantita is None:
            skipped_reasons["no_quantity"] += 1
            continue

        quantita, importo = calculate_line_amount(quantita, prezzo_unitario)

        voce = ParsedItem(
            order=ordine,
            progressive=progressivo,
            code=codice,
            description=descrizione,
            wbs_levels=_extract_wbs_levels(data_row, description_indexes),
            unit=None,
            quantity=quantita,
            unit_price=prezzo_unitario,
            amount=importo,
            notes=None,
            metadata={
                "column_warnings": column_warnings or None,
                "tokens": tokens,
            },
            long_description=descrizione_estesa
        )
        voci.append(voce)
        ordine += 1

    # Log summary
    logger.info(f"[LX Parser] Parsed {len(voci)} items, skipped: {skipped_reasons}")

    total_amount = ceil_amount(sum([voce.amount or 0 for voce in voci])) if voci else None
    return ParsedEstimate(
        title=sheet.title if sheet else None,
        total_amount=total_amount,
        total_quantity=None,
        items=voci,
        stats=None,
    )


def _ensure_indexes_lc(
    name: str,
    columns: Sequence[str],
    data_rows,
    header_row,
    warnings: list[str],
    suggestions: dict,
) -> list[int]:
    try:
        indexes = columns_to_indexes(columns, name, header_row=header_row, required=False)
    except ValueError:
        indexes = []
    if indexes and _has_values(data_rows, indexes):
        return indexes
    suggestion = suggestions.get(name)
    if suggestion:
        warnings.append(
            f"Colonna {name} non trovata, uso automaticamente '{suggestion.header_label or suggestion.column_letter}'."
        )
        return [suggestion.column_index]
    available = ", ".join([cell for cell in header_row if cell]) if header_row else ""
    raise ValueError(f"Impossibile individuare la colonna {name}. Intestazioni rilevate: {available or 'nessuna'}")


def _has_values(rows, indexes: list[int]) -> bool:
    for row in rows:
        for idx in indexes:
            if idx < len(row) and row[idx] not in (None, "", " "):
                return True
    return False


def _cell_to_progressive(row, index: int | None) -> int | None:
    if index is None or index >= len(row):
        return None
    value = row[index]
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _extract_wbs_levels(row, description_indexes: list[int]) -> list[ParsedWbsLevel]:
    levels: list[ParsedWbsLevel] = []
    for idx in description_indexes:
        if idx < len(row):
            value = row[idx]
            if value and isinstance(value, str) and looks_like_wbs7_code(value):
                normalized = normalize_wbs7_code(value)
                if normalized:
                    levels.append(ParsedWbsLevel(level=7, code=normalized, description=None))
                    break
    return levels


__all__ = ["parse_lx_return_excel"]
