from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_excel_workbook(file_path: Path, data_only: bool = True, read_only: bool = True):
    """
    Load an Excel workbook.
    """
    return load_workbook(filename=file_path, data_only=data_only, read_only=read_only)


def select_sheet(workbook, sheet_name: str | None) -> Worksheet:
    """
    Select a specific sheet by name, or the active sheet if name is None or not found.
    """
    if sheet_name and sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.active


def read_all_rows(sheet: Worksheet, values_only: bool = True) -> list[tuple[Any, ...]]:
    """
    Read all rows from a worksheet.
    """
    return list(sheet.iter_rows(values_only=values_only))
