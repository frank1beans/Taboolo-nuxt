from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl.utils import get_column_letter

from parsers.shared.base_types import MAX_WBS_LEVELS
from parsers.shared.math_utils import _calculate_line_amount
from parsers.excel.detector import (
    ColumnSuggestion,
    ColumnProfile,
    detect_column_suggestions,
    ensure_indexes,
    warn_and_use,
    columns_to_indexes,
    single_column_index,
    format_profiles,
)
from parsers.excel.processor import (
    drop_empty_columns,
    locate_header_row,
    column_has_values,
    row_has_values,
    combine_code,
    combine_text,
    cell_to_float,
    cell_to_progressive,
    apply_column_filter,
    extract_wbs_levels,
    sanitize_price_candidate,
)
from parsers.excel.reader import load_excel_workbook, select_sheet, read_all_rows
from parsers.excel.types import ParsedEstimate, ParsedItem
from parsers.helpers.text_and_measure import head_to_tail_quantity, tokenize_description


def parse_estimate_excel(
    file_path: Path,
    sheet_name: str | None = None,
    price_column: str | None = None,
    quantity_column: str | None = None,
) -> ParsedEstimate:
    """
    Generic Excel estimate parser with auto-detection.
    This is a simplified wrapper - for full control use parse_custom_return_excel.
    """
    return parse_custom_return_excel(
        file_path,
        sheet_name,
        code_columns=[],
        description_columns=[],
        price_column=price_column or "",
        quantity_column=quantity_column,
        combine_totals=True,
    )


def parse_custom_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
    *,
    combine_totals: bool = True,
) -> ParsedEstimate:
    """Public API: wrapper around `_parse_custom_return_excel`."""
    return _parse_custom_return_excel(
        file_path,
        sheet_name,
        code_columns,
        description_columns,
        price_column,
        quantity_column,
        progressive_column,
        combine_totals=combine_totals,
    )


def _parse_custom_return_excel(
    file_path: Path,
    sheet_name: str | None,
    code_columns: Sequence[str],
    description_columns: Sequence[str],
    price_column: str,
    quantity_column: str | None = None,
    progressive_column: str | None = None,
    *,
    combine_totals: bool = True,
) -> ParsedEstimate:
    workbook = load_excel_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = select_sheet(workbook, sheet_name)
        raw_rows = read_all_rows(sheet)
    finally:
        workbook.close()

    # Pre-normalize table: drop fully empty columns and fill missing values where needed
    rows, dropped_columns, kept_column_indexes = drop_empty_columns(raw_rows)
    header_idx = locate_header_row(rows)
    if header_idx is None:
        raise ValueError("The selected Excel sheet does not contain importable rows.")
    data_rows = rows[header_idx + 1 :]
    if not data_rows:
        raise ValueError("The selected sheet does not contain data after the header row.")

    header_row = rows[header_idx]
    detection = detect_column_suggestions(rows, header_idx)
    suggestions: dict[str, ColumnSuggestion] = detection.get("suggestions", {})
    profiles: list[ColumnProfile] = detection.get("profiles", [])

    column_warnings: list[str] = []
    if dropped_columns:
        column_warnings.append(
            f"Automatically ignored {dropped_columns} completely empty columns."
        )
    try:
        code_indexes = columns_to_indexes(code_columns, "codice", header_row=header_row, required=False)
    except ValueError:
        code_indexes = []
    try:
        description_indexes = columns_to_indexes(description_columns, "descrizione", header_row=header_row, required=False)
    except ValueError:
        description_indexes = []
    try:
        price_index = single_column_index(price_column, "prezzo unitario", header_row=header_row)
    except ValueError:
        price_index = None
    try:
        quantity_index = single_column_index(quantity_column, "quantita", header_row=header_row, required=False)
    except ValueError:
        quantity_index = None
    try:
        progressive_index = single_column_index(progressive_column, "progressivo", header_row=header_row, required=False)
    except ValueError:
        progressive_index = None

    code_indexes = ensure_indexes("codice", code_indexes, data_rows, suggestions, profiles, column_warnings)
    description_indexes = ensure_indexes("descrizione", description_indexes, data_rows, suggestions, profiles, column_warnings)
    
    if price_index is None:
        suggestion = suggestions.get("prezzo")
        price_index = warn_and_use("prezzo", suggestion, column_warnings)
    if price_index is None:
        available = format_profiles(profiles)
        raise ValueError(
            "Cannot find the unit price column. "
            f"Detected headers: {available or 'no valid headers.'}"
        )
    if not column_has_values(data_rows, [price_index]):
        raise ValueError(
            f"The selected price column ({get_column_letter(price_index + 1)}) is empty. "
            "Check the chosen column in the sheet."
        )
    if quantity_index is None:
        quantity_index = warn_and_use("quantity", suggestions.get("quantita"), column_warnings)
    elif quantity_index is not None and not column_has_values(data_rows, [quantity_index]):
        raise ValueError(
            f"The selected quantity column ({get_column_letter(quantity_index + 1)}) is empty. "
            "Check the chosen column in the sheet."
        )
    if progressive_index is None:
        progressive_index = warn_and_use("progressivo", suggestions.get("progressivo"), column_warnings)
    elif progressive_index is not None and not column_has_values(data_rows, [progressive_index]):
        raise ValueError(
            f"The selected progressive column ({get_column_letter(progressive_index + 1)}) is empty. "
            "Check the chosen column in the sheet."
        )
    if not code_indexes and not description_indexes and progressive_index is None:
        raise ValueError(
            "Select at least one column for code, description, or progressive."
        )
    
    workbook_formulas = load_excel_workbook(file_path, data_only=False, read_only=True)
    try:
        formula_sheet = select_sheet(workbook_formulas, sheet_name)
        # Note: min_row is 1-based in openpyxl, header_idx is 0-based
        raw_formula_rows = list(formula_sheet.iter_rows(min_row=header_idx + 2, values_only=False))
        formula_rows = apply_column_filter(raw_formula_rows, kept_column_indexes)
    finally:
        workbook_formulas.close()

    # formula_rows_iter = iter(formula_rows) # Unused in original code?

    items: list[ParsedItem] = []
    order = 0

    # Head-to-tail logic: 1 progressive = 1 item
    cleaned_rows = []
    for row in data_rows:
        if not row_has_values(row):
            continue

        code_value = combine_code(row, code_indexes)
        description_value = combine_text(row, description_indexes)
        raw_price = cell_to_float(row, price_index)
        quantity_value_head = cell_to_float(row, quantity_index) if quantity_index is not None else None
        progressive_value = cell_to_progressive(row, progressive_index)

        # Skip summary rows only if empty
        if (
            description_value
            and description_value.lower().startswith("totale ")
            and not progressive_value
            and raw_price is None
            and quantity_value_head is None
        ):
            continue

        # Skip chapter/title rows
        if (code_value or description_value) and progressive_value is None and quantity_value_head is None and raw_price is None:
            continue

        cleaned_rows.append(row)

    progressive_indexes = []
    for idx, row in enumerate(cleaned_rows):
        progressive_value = cell_to_progressive(row, progressive_index)
        if progressive_value is not None:
            progressive_indexes.append((idx, progressive_value, row))

    if not progressive_indexes:
        raise ValueError("No progressive value found: select the correct column.")

    for prog_idx, prog_value, row in progressive_indexes:
        description_value = combine_text(row, description_indexes)
        unit_price = sanitize_price_candidate(cell_to_float(row, price_index))

        # Quantity: either from the same row or from following measure rows
        quantity_value = cell_to_float(row, quantity_index) if quantity_index is not None else None
        if quantity_value is None:
            quantity_value = head_to_tail_quantity(cleaned_rows, prog_idx)

        if quantity_value is None:
            # logger_warning = (
            #     f"Progressive {prog_value}: missing quantity, defaulting to 0."
            # )
            # print(logger_warning)
            quantity_value = 0.0

        quantity_value, amount = _calculate_line_amount(quantity_value, unit_price)
        tokens = tokenize_description(description_value or "")

        item = ParsedItem(
            order=order,
            progressive=prog_value,
            code=combine_code(row, code_indexes),
            description=description_value,
            wbs_levels=extract_wbs_levels(row, description_indexes),
            unit=None,
            quantity=quantity_value,
            unit_price=unit_price,
            amount=amount,
            notes=None,
            metadata={
                "sheet_name": sheet.title if sheet else None,
                "column_warnings": column_warnings or None,
                "tokens": tokens,
            },
        )
        items.append(item)
        order += 1

    # If combine_totals=False, skip sum of total rows
    total_amount = sum([item.amount or 0 for item in items]) if combine_totals else None
    return ParsedEstimate(
        title=sheet.title if sheet else None,
        total_amount=total_amount,
        total_quantity=None,
        items=items,
        stats=None,
    )
